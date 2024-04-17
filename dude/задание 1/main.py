import os
import logging
from tkinter import Tk, filedialog, messagebox
from docx2pdf import convert
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

# Настройки логирования
logging.basicConfig(filename='conversion_log.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def convert_docx_to_pdf(docx_file):
    # Конвертирование документа на пдф
    try:
        pdf_path = filedialog.asksaveasfilename(title="Сохранить PDF как", defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if pdf_path:
            convert(docx_file, pdf_path)
            logging.info(f"Конвертация {docx_file} в PDF успешно завершена")
            messagebox.showinfo("Успех", f"PDF успешно сохранен как {pdf_path}")            
    except Exception as e:
        logging.error(f"Ошибка при конвертации {docx_file} в PDF: {str(e)}")
        messagebox.showerror("Ошибка", f"Ошибка при конвертации {docx_file} в PDF: {str(e)}")

def create_pdf_from_xlsx(xlsx_file):
    # Конвертирование таблицу на пдф
    try:
        pdf_path = filedialog.asksaveasfilename(title="Сохранить PDF как", defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if pdf_path:
            wb = load_workbook(xlsx_file)
            pdf = SimpleDocTemplate(pdf_path, pagesize=letter)
            table_data = []
            # Сохранение данных на массив
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    table_data.append(row)
            # Создание стиля таблицы
            table = Table(table_data)
            style = TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), '#4CAF50'),  # Фон заголовка таблицы (зеленый цвет)
                        ('TEXTCOLOR', (0, 0), (-1, 0), (1, 1, 1, 1)),  # Цвет текста в заголовке таблицы (белый цвет)
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Выравнивание текста по центру
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Жирный шрифт для заголовка таблицы
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Отступ снизу у заголовка таблицы
                        ('BACKGROUND', (0, 1), (-1, -1), '#FFEB3B'),  # Фон ячеек таблицы (желтый цвет)
                        ('GRID', (0, 0), (-1, -1), 1, '#000000')  # Цвет и толщина границ таблицы (черный цвет)
                    ])

            table.setStyle(style)
            pdf.build([table])
            logging.info(f"PDF успешно сохранен как {pdf_path}")
            messagebox.showinfo("Успех", f"PDF успешно сохранен как {pdf_path}")
    except Exception as e:
        logging.error(f"Ошибка при создании PDF из {xlsx_file}: {str(e)}")
        messagebox.showerror("Ошибка", f"Ошибка при создании PDF из {xlsx_file}: {str(e)}")

def main():
    # Создание окна для выбора файла
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="Выберите файл для конвертации/обработки")
    # Проверка файла Таблица или Документ
    if file_path:
        if file_path.endswith('.docx'):
            convert_docx_to_pdf(file_path)
        elif file_path.endswith('.xlsx'):
            create_pdf_from_xlsx(file_path)
        else:
            logging.error("Неподдерживаемый формат файла")
            messagebox.showerror("Ошибка", "Неподдерживаемый формат файла")
    else:
        logging.error("Файл не выбран")
        messagebox.showerror("Ошибка", "Файл не выбран")

# Пример использования скрипта
if __name__ == "__main__":
    main()
