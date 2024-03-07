# Импортируем нужные библиотеки
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import logging
from docx2pdf import convert
from openpyxl import load_workbook

# Для начала мы указываем место где будут находиться логи и формат в котором они будут писаться
logging.basicConfig(filename='conversion_logs.txt', level=logging.INFO,
                    format='%(asctime)s:%(levelname)s:%(message)s')


# Первым создадим функцию для конвертации файла из формата docx в pdf
def convert_docx_to_pdf(input_path):
    try:
        output_path = input_path.replace('.docx', '.pdf')  # Мы будем хранить файл pdf, там же где и сам docx файл.
        convert(input_path, output_path)  # Вызываем функцию конвертации, которая сделает все за нас.
        logging.info(f"Converted {input_path} to {output_path}")  # При удачной конвертации, данные будем записывать в логи
        messagebox.showinfo("Успешно", f"Файл {input_path} конвертирован в {output_path}")
        # Так же поскольку работаем с tkinter, можно так же вывести информацию об успешной конвертации в окне.
    except Exception as err:
        messagebox.showerror("Ошибка", str(err))  # Если конвертация не удасться, тогда нам выведет окно ошибки
        logging.error(f"Error occurred when try to convert {input_path} file")
        logging.error(str(err))  # Так же это запишем в логи


# Функция для вывода списка листов в xlsx файле
def list_sheets(input_path):
    try:
        wb = load_workbook(filename=input_path)  # Загружаем книгу Excel для работы с ней
        sheets_info = "\n".join([f"{i+1}: {sheet}" for i, sheet in enumerate(wb.sheetnames)])  # Формируем строку с именами и номерами листов
        logging.info(f"Listed sheets for {input_path}: {sheets_info}")  # Записываем информацию о листах в лог
        messagebox.showinfo("Листы", sheets_info)  # Выводим информацию в окне сообщения
    except Exception as err:
        messagebox.showerror("Ошибка", str(err))  # Выводим окно ошибки, если что-то пошло не так
        logging.error(f"Error listing sheets for {input_path}")  # Записываем ошибку в лог
        logging.error(str(err))


# Функция для вывода всех не пустых строк в выбранном листе
def show_non_empty_rows(input_path):
    sheet_name = simpledialog.askstring("Ввод", "Введите имя листа:")  # Запрашиваем имя листа у пользователя
    if sheet_name:
        try:
            wb = load_workbook(filename=input_path)  # Загружаем книгу Excel
            sheet = wb[sheet_name]  # Получаем лист по имени
            values = []
            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    values.append(", ".join([str(cell) for cell in row if cell is not None]))  # Собираем данные из непустых ячеек
            logging.info(f"Displayed non-empty rows for {sheet_name} in {input_path}")  # Логируем успешное отображение данных
            messagebox.showinfo("Данные листа", "\n".join(values))  # Показываем данные в модальном окне
        except Exception as err:
            messagebox.showerror("Ошибка", str(err))  # Выводим сообщение об ошибке
            logging.error(f"Error displaying non-empty rows for {sheet_name} in {input_path}")  # Записываем ошибку в лог
            logging.error(str(err))


# Функция для добавления тестовых данных в первые три строки xlsx файла
def add_test_data(input_path):
    try:
        wb = load_workbook(filename=input_path)  # Открываем книгу Excel
        sheet = wb.active  # Работаем с активным листом
        values = ['Тест1', 'Тест2', 'Тест3']
        for i, value in enumerate(values, 1):
            sheet[f'A{i}'] = value  # Добавляем тестовые данные в первые три строки столбца A
        wb.save(input_path)  # Сохраняем изменения в файле
        logging.info(f"Test data added to {input_path}")  # Логируем успешное добавление тестовых данных
        messagebox.showinfo("Успешно", "Тестовые данные добавлены")  # Информируем пользователя об успешном добавлении данных
    except Exception as err:
        messagebox.showerror("Ошибка", str(err))  # В случае ошибки выводим сообщение
        logging.error(f"Error adding test data to {input_path}")  # Записываем ошибку в лог
        logging.error(str(err))


# Основная функция для выбора действия и файла
def action_with_file(action):
    file_path = filedialog.askopenfilename()  # Открываем диалог выбора файла
    if file_path:
        if action == "convert":
            convert_docx_to_pdf(file_path)  # Вызываем функцию конвертации для docx в pdf
        elif action == "list_sheets":
            list_sheets(file_path)  # Вызываем функцию для отображения списка листов
        elif action == "show_non_empty":
            show_non_empty_rows(file_path)  # Функция для отображения непустых строк
        elif action == "add_test_data":
            add_test_data(file_path)  # Функция для добавления тестовых данных


# Графический интерфейс пользователя для выбора действия
def main_window():
    root = tk.Tk()
    root.title("Выберите действие")

    # Кнопки для различных действий
    tk.Button(root, text="Конвертировать DOCX в PDF", command=lambda: action_with_file("convert")).pack(fill=tk.X)
    tk.Button(root, text="Показать листы XLSX", command=lambda: action_with_file("list_sheets")).pack(fill=tk.X)
    tk.Button(root, text="Показать не пустые строки", command=lambda: action_with_file("show_non_empty")).pack(fill=tk.X)
    tk.Button(root, text="Добавить тестовые данные", command=lambda: action_with_file("add_test_data")).pack(fill=tk.X)

    root.mainloop()  # Запускаем главное окно приложения


if __name__ == "__main__":
    main_window()  # Вызов главной функции для старта приложения

