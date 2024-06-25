import os
import sys
import logging
import docx2pdf
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from tkinter import messagebox
from docx2pdf import convert
from openpyxl import load_workbook

# Найстройка логирования
logging.basicConfig(filename='logs.txt', level=logging.INFO,
                    format='%(asctime)s: %(levelname)s: %(message)s')


def log_error(err):
    logging.error(err)
    messagebox.showerror('Ошибка', err) # Вывод ошибки в модальном окне


def convert_docx(docx_path):
    try:
        pdf_path = os.path.splitext(docx_path)[0] + '.pdf'
        convert(docx_path, pdf_path)
        logging.info(f'Успешно конвертировано {docx_path} в {pdf_path}.')
    except Exception as e:
        log_error(f'Не удалось конвертировать {docx_path}. Ошибка: {str(e)}')


def convert_xlsx(excel_path):
    try:
        wb = load_workbook(excel_path) # загрузка Excel файла
        sheetnames = wb.sheetnames # получение листов Excel файла
        pdf_path = os.path.splitext(excel_path)[0] + '.pdf'
        pp = PdfPages(pdf_path)
        for sheet in sheetnames:
            df = pd.read_excel(excel_path, sheet_name=sheet) # преобразование листа в dataframe
            fig, ax = plt.subplots(figsize=(10, 10))
            ax.axis('tight')
            ax.axis('off')
            # рисование таблицы Excel с использованием matplot
            the_table = ax.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')
            the_table.scale(1.5, 1.5)
            plt.title(sheet, fontsize=12)
            # сохранение рисунка таблицы в pdf файле
            pp.savefig(fig, bbox_inches='tight')
            plt.close(fig)
        pp.close()
        logging.info(f'Успешно обработан {excel_path} и сохранен в {pdf_path}.')
    except Exception as e:
        log_error(f'Не удалось конвертировать {excel_path}. Ошибка: {str(e)}')


def main(file_path):
    file_extension = os.path.splitext(file_path)[1]
    # проверка корректности формата файла
    if file_extension.lower() == '.docx':
        convert_docx(file_path)
    elif file_extension.lower() == '.xlsx':
        convert_xlsx(file_path)
    else:
        log_error('Формат файла не поддерживается. Выберите форматы .xlsx, .docx')


if __name__ == '__main__':
    if len(sys.argv) != 2:
        log_error('Ошибка в вызове скрипта. Попробуйте: python main.py <путь_к_файлу>')
    else:
        file_path = sys.argv[1]
        if os.path.exists(file_path):
            main(file_path)
        else:
            log_error(f'Файл не найден: {file_path}')



